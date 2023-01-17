"""
Contains utilitarian methods/functionalities/data
"""
from builtins import staticmethod

import numpy
import openpyxl
import copy
import os
import re
import hashlib
import time
import multiprocessing
import multiprocessing.shared_memory
import pathlib
import xlrd
import json
import collections
import importlib
import operator
import gc
import codecs
import itertools
import sklearn.metrics
import datetime
import subprocess
import warnings
import sklearn.exceptions
import joblib
import functools
import shutil
import matplotlib.pyplot as plt
import contextlib
import inspect

import propheticus

if hasattr(propheticus.Config, 'framework_selected_instance_path'):
    import sys
    sys.path.append(propheticus.Config.framework_selected_instance_path)
    if os.path.isfile(os.path.join(propheticus.Config.framework_selected_instance_path, 'InstanceConfig.py')):
        from InstanceConfig import InstanceConfig as Config
    else:
        import propheticus.Config as Config
else:
    import propheticus.Config as Config

def static_vars(**kwargs):
    def decorate(func):
        for k in kwargs:
            setattr(func, k, kwargs[k])
        return func

    return decorate


class Utils:
    THREAD_LEVEL_BATCH = 'batch'
    THREAD_LEVEL_RUN = 'run'
    THREAD_LEVEL_CV = 'cv'
    THREAD_LEVEL_ALGORITHM = 'algorithm'
    THREAD_LEVEL_BEST = 'best'

    PYTHON_NUMPY_DATA_TYPES_MAP = {
        'int64': numpy.int64,
        'float64': numpy.float64,
        'string': numpy.object,
        'bool': numpy.bool
    }

    ColourValues = [
                     'FF0000', '00FF00', '0000FF', 'FFFF00', '00FFFF', 'FF00FF', '000000',
                     '800000', '008000', '000080', '808000', '800080', '008080', '808080',
                     'C00000', '00C000', '0000C0', 'C0C000', 'C000C0', '00C0C0', 'C0C0C0',
                     '400000', '004000', '000040', '404000', '400040', '004040', '404040',
                     '200000', '002000', '000020', '202000', '200020', '002020', '202020',
                     '600000', '006000', '000060', '606000', '600060', '006060', '606060',
                     'A00000', '00A000', '0000A0', 'A0A000', 'A000A0', '00A0A0', 'A0A0A0',
                     'E00000', '00E000', '0000E0', 'E0E000', 'E000E0', '00E0E0', 'E0E0E0',
                    ]

    AvailableExperiments = {True: None, False: None}

    LoggedMessages = {'error': [], 'status': [], 'warning': []}

    AvailableDataBalancing = sorted([algorithm for algorithm, Details in Config.SamplingCallDetails.items()])
    AvailableDimensionalityReduction = sorted([algorithm for algorithm, Details in Config.DimensionalityReductionCallDetails.items()])
    AvailableClassificationAlgorithms = sorted([algorithm for algorithm, Details in Config.ClassificationAlgorithmsCallDetails.items()])
    AvailableClusteringAlgorithms = sorted([algorithm for algorithm, Details in Config.ClusteringAlgorithmsCallDetails.items()])

    # TODO: this needs to be more flexible
    AvailableClassificationMetrics = sorted(['f05-score', 'f1-score', 'f2-score', 'roc-auc', 'accuracy', 'precision', 'recall', 'specificity', 'informedness', 'markedness'])

    RandomSeeds = [341187527, 644570282, 786112152, 135708657, 597861104, 880607938, 787554810, 311771549, 717616322, 276791663,
                   852934353, 562682195, 881244661, 954295709, 911889983, 69598916, 359584179, 135103005, 39271315, 137908699,
                   142079615, 448358036, 238576428, 295203401, 170362978, 341448958, 701279747, 914483372, 834810093, 115986650,
                   789409302, 142402517, 687469447, 334410270, 711362491, 407710244, 431944790, 785539881, 608282410, 251431605,
                   649328852, 892069364, 344664614, 762394826, 930487039, 989782233, 593976005, 239724272, 497825617, 462618158,
                   844769795, 610829263, 211491405, 833417652, 367364232, 278647467, 173341764, 327047894, 293761972, 216416693,
                   731775212, 351181789, 307283204, 310647995, 785398587, 648291807, 589279062, 342193879, 627056663, 687649406,
                   665023227, 841534990, 838455756, 500297361, 475517064, 954818250, 770296613, 965471622, 398167053, 177041683,
                   947366292, 679777755, 460266078, 835126884, 266543261, 941582404, 983662003, 199972135, 492276935, 246620637,
                   221994739, 561959825, 669789814, 228404659, 708026157, 657738783, 913368682, 499376167, 551579025, 450908424,
                   940996345, 646799793, 964924459, 600475941, 935482053, 570167920, 660060394, 537006218, 597534927, 900337911,
                   361438669, 435852050, 242948452, 914840717, 638272806, 694694184, 725383302, 829455291, 543116649, 608233042,
                   540947413, 983970825, 711093705, 702656598, 429432983, 927060905, 629295323, 834723964, 911742988, 643573763,
                   727181671, 990681549, 718957436, 698302669, 333021048, 486422386, 866820962, 205792901, 147279871, 504639404,
                   538879304, 728579439, 941378653, 534156448, 513382453, 187569883, 297741642, 517597085, 609291501, 216640123,
                   956231391, 948409093, 374938828, 172807776, 847296497, 371245439, 342112431, 460888728, 771509136, 452786562,
                   622926861, 398174344, 509021461, 198267415, 951490814, 559714656, 431878120, 530723896, 118684391, 701552706,
                   774394751, 537640559, 414923699, 427658427, 867564946, 101635231, 358524192, 240059507, 760356253, 382429946,
                   709919973, 579669379, 984578969, 315486587, 395545234, 863568195, 921766771, 675797641, 828418817, 482890883,
                   878993063, 892780593, 106564210, 334102057, 108042673, 472494095, 743850318, 758284763, 515961726, 205368468,
                   633660232, 545108815, 295249597, 812233851, 430433126, 684627199, 313178714, 263356970, 689223497, 369919880,
                   461197301, 148246000, 749559795, 802644312, 775077493, 899026598, 533222372, 796779498, 730745312, 873665382,
                   124126932, 287803450, 283046372, 587345018, 942264457, 673284996, 924400620, 346110134, 513398186, 297635124]

    hide_demo_popups = False
    cachedStep = None
    chars_per_second = 15

    '''
    Data Choices for the UI Related Functions
    '''
    @staticmethod
    def getPersistedModelsExperimentsList():
        return [Experiment['label'] for experiment_identifier, Experiment in Utils.getPersistedModelsExperiments().items()]

    @staticmethod
    def getPersistedModelsExperiments():
        if not os.path.isdir(Config.framework_instance_generated_persistent_path):
            return {}

        AvailableExperiments = Utils.getAvailableExperiments()

        Experiments = {}
        WarningExperiment = {}
        for filename in os.listdir(Config.framework_instance_generated_persistent_path):
            if '~' not in filename:
                experiment_identifier = filename.split('.')[0]
                experiment_base_name = filename.split('-')[0]
                if experiment_base_name in Experiments:
                    continue

                if experiment_identifier in AvailableExperiments:
                    Experiments[experiment_base_name] = copy.deepcopy(AvailableExperiments[experiment_identifier])
                    Experiments[experiment_base_name]['label'] = f'{experiment_base_name} - ' + ' | '.join(map(str, Experiments[experiment_base_name]['brief_details'].values()))
                elif experiment_base_name not in WarningExperiment:
                    WarningExperiment[experiment_base_name] = True
                    Utils.printWarningMessage(f'Persisted models found but no corresponding log was found: {experiment_base_name}')

        return collections.OrderedDict(sorted(Experiments.items()))

    @staticmethod
    def getAvailableExperimentsList(use_cached=True, skip_config_parse=False, field='label'):
        return [Experiment[field] for experiment_identifier, Experiment in Utils.getAvailableExperiments(use_cached, skip_config_parse).items()]

    @staticmethod
    def getExperimentDetailsByFile(path, filename):
        experiment_identifier = filename.split('.')[0]

        Workbook = openpyxl.load_workbook(os.path.join(path, filename))
        DetailsWorksheet = Workbook.worksheets[0]
        experiment_configuration = DetailsWorksheet.cell(1, 1).value
        ParsedConfigurations = Utils.parseLogStringConfiguration(experiment_configuration)

        ExperimentDetails = {
            'identifier': experiment_identifier,
            'path': path,
            'configuration': ParsedConfigurations,
            'filename': filename,
        }

        return ExperimentDetails

    @staticmethod
    def getExperimentDetailsById(experiment_id):
        for root, SubFolders, Files in os.walk(os.path.join(Config.framework_instance_generated_logs_path)):
            for subfolder in SubFolders:
                for filename in os.listdir(os.path.join(root, subfolder)):
                    if experiment_id in filename and '.Log.' in filename and '~' not in filename:
                        return Utils.getExperimentDetailsByFile(os.path.join(root, subfolder), filename)

    @staticmethod
    def _getAvailableExperiments(root, subfolder, filename, filter_by):
        if '~' in filename or '.Log.' not in filename:
            return None

        DescriptiveFields = [
            'pre_target',
            'config_ensemble_algorithms',
            'proc_balance_data',
            'proc_classification',
            'proc_classification_algorithms_parameters',
            'proc_reduce_dimensionality',
        ]

        experiment_identifier = filename.split('.')[0]

        ExperimentDetails = Utils.getExperimentDetailsByFile(os.path.join(root, subfolder), filename)
        ParsedConfigurations = ExperimentDetails['configuration']

        BriefDetails = {key: value for key, value in ExperimentDetails['configuration'].items() if key in DescriptiveFields}

        if filter_by is not None:
            if 'proc_' + filter_by not in ParsedConfigurations or not ParsedConfigurations['proc_' + filter_by]:
                return None

        ExperimentDetails = {
            'identifier': experiment_identifier,
            'label': ' ' + subfolder + ' >> ' + ' | '.join(map(str, BriefDetails.values())) + ' * ' + filename,
            'brief_details': BriefDetails,
            'configuration': ParsedConfigurations,
            'filename': filename,
            'subfolder': subfolder
        }

        return ExperimentDetails

    @staticmethod
    def getAvailableExperiments(use_cached=True, skip_config_parse=False, filter_by=None):
        if Utils.AvailableExperiments[skip_config_parse] is None or use_cached is False:
            Experiments = {}
            for root, SubFolders, Files in os.walk(os.path.join(Config.framework_instance_generated_logs_path)):
                for subfolder in SubFolders:
                    Files = os.listdir(os.path.join(root, subfolder))
                    if not skip_config_parse:
                        pool_count = min(Config.max_thread_count, len(Files))

                        if not Utils.isMultiProcess():
                            FolderExperiments = propheticus.shared.Utils.pool(pool_count, Utils._getAvailableExperiments, [(root, subfolder, filename, filter_by) for filename in Files])
                        else:
                            FolderExperiments = [Utils._getAvailableExperiments(root, subfolder, filename, filter_by) for filename in Files]

                        for Experiment in FolderExperiments:
                            if Experiment is not None:
                                Experiments[Experiment['identifier']] = Experiment

                    else:
                        for filename in Files:
                            if '~' not in filename and '.Log.' in filename:
                                experiment_identifier = filename.split('.')[0]
                                Experiments[experiment_identifier] = {
                                    'identifier': experiment_identifier,
                                    'label': ' ' + subfolder + ' >> * ' + filename,
                                    'brief_details': {},
                                    'configuration': {},
                                    'filename': filename,
                                    'subfolder': subfolder
                                }

            Utils.AvailableExperiments[skip_config_parse] = collections.OrderedDict(sorted(Experiments.items()))

        return Utils.AvailableExperiments[skip_config_parse]

    @staticmethod
    def getAvailableDatasets():
        return sorted([file for file in os.listdir(Config.framework_instance_data_path) if '.data.txt' in file])

    ''''###############################################################################################'''



    '''
    Utility Related Functions
    '''
    @staticmethod
    def deleteFromStructuredArray(a, column):
        ''' Delete columns from structured array '''
        return a[[name for name in a.dtype.names if name not in column]]

    @staticmethod
    def multipleReplace(str, Replace):
        for _find, _replace in Replace:
            str = str.replace(_find, _replace)

        return str.strip()

    @staticmethod
    def loadExperienceModels(experiment_base_name, persistent_path=None, Estimators=None):
        if persistent_path is None:
            persistent_path = Config.framework_instance_generated_persistent_path

        if Estimators is None:
            Estimators = ['variance', 'normalize', 'reduce_dimensionality', 'balance_data', 'classifier']

        Transformers = {}
        for item in os.listdir(persistent_path):
            if experiment_base_name in item:
                if '.headers.' in item:
                    with open(os.path.join(persistent_path, item)) as f:
                        Transformers['headers'] = [line for line in f.read().split('\n') if line.strip() != '']
                else:
                    found_estimator = propheticus.shared.Utils.inString(item, Estimators)
                    if found_estimator is False:
                        continue

                    Transformers[found_estimator] = propheticus.shared.Utils.loadModelFromDist(persistent_path, item)

        return Transformers

    @staticmethod
    def removeLibSVMSparsity(file, num_features, zero_based):
        with open(file) as f:
            TestLines = f.read().split('\n')

        ParsedLines = []
        for index, line in enumerate(TestLines):
            if line.strip() == '':
                continue

            Details = line.split()
            if len(Details) != num_features + 1:
                index_start = 0 if zero_based is True else 1
                label_offset = 1 if zero_based is True else 0
                index_end = num_features if zero_based is True else num_features + 1
                for i in range(index_start, index_end):
                    if f' {i}:' not in line:  # NOTE: keep the space
                        Details.insert(i + label_offset, f'{i}:0.0')  # i + 1 to keep the label at 0
                        # print(f'Inserting feature value {i} at line {index}')

                ParsedLines.append(' '.join(Details))
            else:
                ParsedLines.append(line)

        with open(file, 'w+') as f:
            f.write('\n'.join(ParsedLines) + '\n')  # NOTE: the last line must be empty, otherwise the last sample is not properly recognized


    @staticmethod
    @contextlib.contextmanager
    def suppress_stdout_stderr():
        """A context manager that redirects stdout and stderr to devnull"""
        with open(os.devnull, 'w') as fnull:
            with contextlib.redirect_stderr(fnull) as err, contextlib.redirect_stdout(fnull) as out:
                yield (err, out)

    @staticmethod
    def moveFilesinDir(srcDir, dstDir):
        pathlib.Path(dstDir).mkdir(parents=True, exist_ok=True)

        if os.path.isdir(srcDir) and os.path.isdir(dstDir):
            # Iterate over all the files in source directory
            for filename in os.listdir(srcDir):
                shutil.move(os.path.join(srcDir, filename), dstDir)
        else:
            print("srcDir & dstDir should be Directories")

    @staticmethod
    def removeColumnsFromList(List, Columns):
        start = time.time()
        if len(List) == 0 or len(Columns) == 0:
            propheticus.shared.Utils.printFatalMessage(f'Columns to be removed and data cannot be empty! {len(List)} - {len(Columns)}')

        one_dimension = False

        max_column = max(Columns)
        list_shape = numpy.array(List).shape
        if len(list_shape) == 1:
            for first_level_index, first_level in enumerate(List):
                if not isinstance(first_level, (list, tuple)):
                    if one_dimension is False and first_level_index != 0:
                        propheticus.shared.Utils.printFatalMessage(f'First level must all be either lists or non-lists')
                    else:
                        one_dimension = True

                if isinstance(first_level, (list, tuple)) and one_dimension is True:
                    propheticus.shared.Utils.printFatalMessage(f'First level must all be either lists or non-lists')

            ''' NOTE
            I changed this function to avoid unexpected behaviours from handling rows with varying length (ie .shape = (X,), and rows that are lists).
            If rows have varying lenght, data will be lost when transposing the data; if necessary, objectively deal it with this situation
            http://code.activestate.com/recipes/410687-transposing-a-list-of-lists-with-different-lengths/
            '''
            if one_dimension is False:
                propheticus.shared.Utils.printFatalMessage(f'Not all rows have the same lenght! {list_shape}')

        if min(Columns) < 0:
            propheticus.shared.Utils.printFatalMessage(f'The lowest possible value is 0: {min(Columns)}')

        if one_dimension is True:
            if max_column >= len(List):
                propheticus.shared.Utils.printFatalMessage(f'The max column value is {len(List) - 1}: {max_column}')

            return [row for index, row in enumerate(List) if index not in Columns]
        else:
            if max_column >= len(List[0]):
                propheticus.shared.Utils.printFatalMessage(f'The max column value is {len(List[0]) - 1}: {max_column}')

            # NOTE: this also works but is slower; however numpy cannot delete from structured arrays
            # ModifiedList = [row for index, row in enumerate(zip(*List)) if index not in Columns]
            # return list(zip(*ModifiedList))

            return numpy.delete(List, Columns, 1).tolist()

    @staticmethod
    def entropy(*X):
        # Source from here: https://blog.biolab.si/2012/06/15/computing-joint-entropy-in-python/
        entropy = numpy.sum(-p * numpy.log2(p) if p > 0 else 0 for p in
            (numpy.mean(functools.reduce(numpy.logical_and, (predictions == c for predictions, c in zip(X, classes))))
                for classes in itertools.product(*[set(x) for x in X])))

        return entropy

    @staticmethod
    def inString(string, match, case_sensitive=True):
        if case_sensitive is False:
            string = string.lower()
            match = match.lower() if isinstance(match, str) else [partial.lower() for partial in match]

        if isinstance(match, str):
            found = match in string
        else:
            Matches = [partial in string for partial in match]
            if True in Matches:
                found = match[Matches.index(True)]
            else:
                found = False

        return found

    @staticmethod
    def saveModelToDisk(path, filename, model):
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

        file_path = os.path.join(path, filename)
        joblib.dump(model, file_path)

        if not os.path.isfile(file_path):
            Utils.printFatalMessage(f'Model was not saved! File: {file_path}')


    @staticmethod
    def loadModelFromDist(path, filename):
        file_path = os.path.join(path, filename)
        if not os.path.isfile(file_path):
            Utils.printFatalMessage(f'The model you tried to load does not exist in the given location: {file_path}')
        else:
            model = joblib.load(file_path)
            return model

    @staticmethod
    def setNestedDictionaryValue(dic, keys, value):
        if isinstance(keys, str):
            keys = [keys]

        for key in keys[:-1]:
            dic = dic.setdefault(key, {})
        dic[keys[-1]] = value

    @staticmethod
    def getNestedDictionaryValue(dic, keys):
        if isinstance(keys, str):
            keys = [keys]

        for key in keys[:-1]:
            dic = dic.setdefault(key, {})
        return dic[keys[-1]]

    @staticmethod
    def temporaryLogs(*args):
        pathlib.Path(Config.framework_temp_path).mkdir(parents=True, exist_ok=True)
        with codecs.open(os.path.join(Config.framework_temp_path, 'execution_log.txt'), "w", encoding="utf-8") as File:
            File.writelines(json.dumps(Utils._getSafeConfigurationsDict(args), indent=2))

    @staticmethod
    def showImage():
        if Config.demonstration_mode is True:
            def close_event():
                Utils.parseDemoStep(prepend='figure')
                time.sleep(2)
                plt.close()

            fig = plt.gcf()
            timer = fig.canvas.new_timer(interval=1500)  # creating a timer object and setting an interval of 3000 milliseconds
            timer.add_callback(close_event)
            timer.start()

        plt.show()

    @staticmethod
    def saveImage(path, filename, **kwargs):
        pathlib.Path(os.path.join(Config.OS_PATH, path)).mkdir(parents=True, exist_ok=True)
        kwargs['fname'] = fname = os.path.join(Config.OS_PATH, path, filename)
        kwargs['transparent'] = Config.use_transparency
        plt.savefig(**kwargs)

        if not os.path.isfile(fname):
            Utils.printFatalMessage('Image was not created! File: ' + fname)

    @staticmethod
    def saveExcel(path, filename, *args, SheetNames=False, show_demo=True):
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

        if SheetNames is False:
            SheetNames = ['Sheet ' + str(index) for index in range(len(args))]

        Workbook = openpyxl.Workbook()
        Sheets = [Workbook.create_sheet(SheetNames[i]) for i in range(1, len(args))]
        Workbook.active.title = SheetNames[0]

        Sheets.insert(0, Workbook.active)

        for index, Data in enumerate(args):
            for Item in Data:
                Sheets[index].append(Item)

        file_path = os.path.join(path, filename)
        Workbook.save(file_path)
        if not os.path.isfile(file_path):
            Utils.printFatalMessage('Excel document was not created! File: ' + file_path)

        if Config.demonstration_mode is True and show_demo and Utils.hide_demo_popups is False:
            Utils.openExternalFileDemo(file_path)

    @staticmethod
    def openExternalFileDemo(file_path):
        with subprocess.Popen(["start", "/WAIT", file_path], shell=True) as doc:
            # use 'doc' here just as you would the file itself
            doc.poll()
            Utils.parseDemoStep(prepend='excel')

    @staticmethod
    def getTimeDifference(start, end=None, precision=3):
        if end is None:
            end = time.time()
        return round((end - start), precision)

    @staticmethod
    def getClassDistribution(Y):
        Y = numpy.array(Y)
        Classes = []
        for index, target in enumerate(Y):
            if target not in Classes:
                Classes.append(target)
        DistributionByClass = {target: int(Y.tolist().count(target)) for target in Classes}

        return DistributionByClass

    @staticmethod
    def getOS():
        if os.name == 'nt':
            system = 'windows'
        elif os.name == 'posix':
            system = 'linux'
        else:
            Utils.printFatalMessage('Unexpected operating system. Change to handle: ' + os.name)

        return system

    @staticmethod
    def getDatasetsIdentifiers(identifier):
        Abrv = {'label_': '', '_': '', 'set_': '', 'XEN': 'X', 'Tomcat': 'Tc', 'vSphere': 'V', 'real': 'R', 'Training': 'Tr', 'Test': 'Ts'}
        for search, replace in Abrv.items():
            identifier = identifier.replace(search, replace)

        if len(identifier) >= 100:
            Utils.printWarningMessage('This would be an hash!')

        return identifier

    @staticmethod
    def validateCurrentUserUUID():
        """
        Validates the current user UUID against the authorized ones defined in the list Config.ValidUUID list.
        WARNING: this varies according to the python version. Python 3.6 has method check_output
        TODO: improve logic
        Returns
        -------

        """
        if Utils.getOS() == 'linux':
            message = subprocess.Popen('dmidecode|grep UUID', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE).communicate()
            if 'denied' in str(message[1]).lower():
                Utils.printFatalMessage('This application must be executed with root privileges')

            uuid = message[0].decode("utf-8").split(' ')[1].split('\n')[0]
            if uuid not in Config.ValidUUID:
                Utils.printFatalMessage('You do not have permission to use this application')
            else:
                Utils.printStatusMessage('User successfully authenticated')
        elif Utils.getOS() == 'windows':
            uuid = subprocess.Popen('wmic csproduct get uuid', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE).communicate()[0].decode("utf-8").split('\n')[1].split(' ')[0]
            # uuid = subprocess.check_output('wmic csproduct get uuid').decode("utf-8").split('\n')[1].strip()
            if uuid not in Config.ValidUUID:
                Utils.printFatalMessage('You do not have permission to use this application')
            else:
                Utils.printStatusMessage('User successfully authenticated')
        else:
            Utils.printFatalMessage('Authentication required and operating system is not supported')

    @staticmethod
    def parseConfusionMatrix(confusion_matrix):
        TruePositiveByClass = {}
        FalsePositivesByClass = {}
        TrueNegativeByClass = {}
        FalseNegativeByClass = {}
        TruePositive = numpy.diag(confusion_matrix)
        FalsePositive = []
        FalseNegative = []
        TrueNegative = []

        num_classes = len(confusion_matrix)
        for i in range(num_classes):
            # NOTE: True Positives
            TruePositiveByClass[i] = TruePositive[i]

            # NOTE: False Positives
            FalsePositivesByClass[i] = sum(confusion_matrix[:, i]) - confusion_matrix[i, i]
            FalsePositive.append(sum(confusion_matrix[:, i]) - confusion_matrix[i, i])

            # NOTE: False Negatives
            FalseNegativeByClass[i] = sum(confusion_matrix[i, :]) - confusion_matrix[i, i]
            FalseNegative.append(sum(confusion_matrix[i, :]) - confusion_matrix[i, i])

            # NOTE: True Negatives
            temp = numpy.delete(confusion_matrix, i, 0)  # delete ith row
            temp = numpy.delete(temp, i, 1)  # delete ith column
            TrueNegativeByClass[i] = sum(sum(temp))
            TrueNegative.append(sum(sum(temp)))

        return TrueNegativeByClass, FalsePositivesByClass, FalseNegativeByClass, TruePositiveByClass


    @staticmethod
    def computeMetricsByClass(Y_true, Y_pred, target_metric=None, average='binary', labels=None):
        # TODO: cache results?

        AllowedAverages = ['weighted', 'macro']
        if average is not None and average not in AllowedAverages:
            propheticus.shared.Utils.printFatalMessage('Allowed averages: ' + ', '.join(AllowedAverages))

        MetricsByClass = {'precision': {}, 'recall': {}, 'informedness': {}, 'markedness': {}, 'specificity': {}}

        confusion_matrix = sklearn.metrics.confusion_matrix(Y_true, Y_pred)
        if confusion_matrix.shape == (1, 1):
            # Utils.printWarningMessage(f'Y_true had only one class and Y_pred correctly predicted all! setting to 0, but value is invalid')
            return None


        TNC, FPC, FNC, TPC = Utils.parseConfusionMatrix(confusion_matrix)
        end = time.time()

        ClassSupport = collections.Counter(sorted(Y_true))
        MissingClasses = sorted(set(Y_pred) - set(Y_true))
        if len(MissingClasses) > 0:
            # propheticus.shared.Utils.printLogMessage(f'Y_pred contains classes not present in Y_true: {MissingClasses}')
            for missing_class in MissingClasses:
                ClassSupport[missing_class] = 0

            ClassSupport = collections.OrderedDict(sorted(ClassSupport.items()))

        for i in range(len(ClassSupport)):
            class_label = list(ClassSupport.keys())[i]
            if TPC[i] + FNC[i] != ClassSupport[class_label]:
                Utils.printFatalMessage(f'The number of samples returned by the confusion matrix does not match the one in ClassSupport!: {i}/{class_label}/{TPC[i] + FNC[i]}/{ClassSupport[class_label]}')

        for i in range(len(ClassSupport)):
            if TPC[i] > 0:
                class_recall = TPC[i] / (TPC[i] + FNC[i])
                class_precision = TPC[i] / (TPC[i] + FPC[i])
            else:
                class_recall = 0
                class_precision = 0

            if TNC[i] > 0:
                class_inverse_recall = TNC[i] / (FPC[i] + TNC[i])
                class_inverse_precision = TNC[i] / (FNC[i] + TNC[i])
            else:
                class_inverse_recall = 0
                class_inverse_precision = 0

            class_label = list(ClassSupport.keys())[i]
            MetricsByClass['precision'][class_label] = class_precision
            MetricsByClass['recall'][class_label] = class_recall
            MetricsByClass['informedness'][class_label] = class_recall + class_inverse_recall - 1
            MetricsByClass['markedness'][class_label] = class_precision + class_inverse_precision - 1
            MetricsByClass['specificity'][class_label] = class_inverse_recall

        if labels is None:
            labels = list(ClassSupport.keys())

        NPositive = sum([ClassSupport[label] for label in labels])
        ComputedMetrics = {}
        if average is not None:
            for metric in MetricsByClass.keys():
                if average == 'weighted':
                    ComputedMetrics[metric] = sum([(ClassSupport[label] / NPositive) * MetricsByClass[metric][label] for label in labels])
                elif average == 'macro':
                    ComputedMetrics[metric] = sum([MetricsByClass[metric][label] for label in labels]) / len(labels)

            with warnings.catch_warnings():
                warnings.filterwarnings(action='ignore', category=sklearn.exceptions.UndefinedMetricWarning)

                # NOTE: for validation purposes
                sklearn_precision = sklearn.metrics.precision_score(Y_true, Y_pred, average=average, labels=labels)
                # sklearn_precision_report = sklearn.metrics.classification_report(Y_true, Y_pred)
                sklearn_recall = sklearn.metrics.recall_score(Y_true, Y_pred, average=average, labels=labels)

            sklearn_precision = round(sklearn_precision, 4)
            sklearn_recall = round(sklearn_recall, 4)
            precision = round(ComputedMetrics['precision'], 4)
            recall = round(ComputedMetrics['recall'], 4)
            tolerance = 0.01

            if precision + tolerance < sklearn_precision or sklearn_precision < precision - tolerance \
                    or recall + tolerance < sklearn_recall or sklearn_recall < recall - tolerance:
                Utils.printFatalMessage(f'Invalid performance computations: {sklearn_precision}:{precision}, {sklearn_recall}:{recall}')
        else:
            ComputedMetrics = MetricsByClass

        return ComputedMetrics if target_metric is None else ComputedMetrics[target_metric]

    @staticmethod
    def cartesianProductDictionaryLists(**kwargs):
        keys, values = zip(*kwargs.items())
        Configs = [dict(zip(keys, v)) for v in itertools.product(*values)]

        return Configs

    @staticmethod
    def isInt(value):
        try:
            int(value)
            return True
        except ValueError:
            return False

    @staticmethod
    def escapeExcelValues(value):
        if isinstance(value, (list, numpy.ndarray, dict)):
            value = str(value)

        return value

    @staticmethod
    def toTitle(value):
        Replace = ['_', '-']
        for search in Replace:
            value = value.replace(search, ' ')

        return value.title()

    @staticmethod
    def getClassDescriptionById(class_id):
        return class_id if class_id not in Config.ClassesDescription else Config.ClassesDescription[class_id]

    @staticmethod
    def getClassIdByDescription(class_description):
        if class_description not in Config.ClassesDescription.values():
            return class_description
        else:
            index = list(Config.ClassesDescription.values()).index(class_description)
            return list(Config.ClassesDescription.keys())[index]

    @staticmethod
    def isMultiProcess():
        return False if multiprocessing.current_process().name == 'MainProcess' else True

    ''''###############################################################################################'''


    '''
    Demonstration Related Functions
    '''
    @staticmethod
    def storeDemonstrationVariables():
        if len(Utils.DemonstrationLogs) > 0:
            demonstration_start_time = datetime.datetime.utcfromtimestamp(Utils.demonstration_start).strftime('%Y%m%d%H%M%S')

            timestamp = propheticus.shared.Utils.getTimeDifference(Utils.demonstration_start, precision=2)
            Utils.DemonstrationLogs[-1][1] = timestamp - Utils.DemonstrationLogs[-1][0]

            propheticus.shared.Utils.saveExcel(
                Config.framework_instance_generated_logs_path,
                demonstration_start_time + '.demonstration_log.xlsx',
                Utils.DemonstrationLogs
            )

            propheticus.shared.Utils.printStatusMessage('Demonstration log successfully saved')

        RemainingSteps = []
        for RemainingStep in Utils.DemonstrationSteps:
            RemainingSteps.append(', '.join(map(str, RemainingStep)))

        if RemainingSteps:
            Utils.printWarningMessage(['There are still steps remaining in the excel file!'] + RemainingSteps)

    @staticmethod
    def cacheNextDemoStep():
        Step = next(Utils.DemonstrationSteps, None)
        if Step is not None:
            Utils.cachedStep = Step

            delay = 0
            Legends = list(map(str, Step[2:]))
            for index, legend in enumerate(Legends):
                if legend is None or legend == 'None':
                    break

                delay += Utils.getSleepDurationByLegend(legend)

            return delay

        else:
            return None

    @staticmethod
    def getSleepDurationByLegend(legend):
        return max(1, len(legend) / Utils.chars_per_second)

    @staticmethod
    def parseDemoStep(prepend=''):
        if Utils.cachedStep is not None:
            Step = Utils.cachedStep
            Utils.cachedStep = None
        else:
            Step = next(Utils.DemonstrationSteps, None)

        if Step is not None:
            action_input = str(Step[0]) if Step[0] is not None else ''
            if prepend.strip() != '' and action_input.strip() != '':
                Utils.printWarningMessage('Legend step with input defined!')

            label = Step[1]
            Legends = list(map(str, Step[2:]))
            for index, legend in enumerate(Legends):
                if legend is None or legend == 'None':
                    if index == 0:
                        Utils.temporaryDemonstrationLogs(f'{label} {prepend}', legend)  # NOTE: REGISTER ONE ENTRY FOR LOG
                    break

                Utils.temporaryDemonstrationLogs(f'{label} {prepend}', legend, sleep_duration=Utils.getSleepDurationByLegend(legend))
        else:
            action_input = label = Legends = None

        return action_input, label, Legends, Step

    @staticmethod
    def initializeDemonstrationVariables():
        Utils.demonstration_start = time.time()
        Utils.DemonstrationLogs = []
        Utils.DemonstrationSteps = []

        Workbook = openpyxl.load_workbook(Config.framework_demo_file_path)
        StepsWorksheet = Workbook.worksheets[0]
        for i in range(2, StepsWorksheet.max_row + 1):
            Utils.DemonstrationSteps.append([StepsWorksheet.cell(i, j).value for j in range(1, StepsWorksheet.max_column + 1)])

        temp = Utils.DemonstrationSteps
        Utils.DemonstrationSteps = iter(Utils.DemonstrationSteps)

        Utils.parseDemoStep()

    @staticmethod
    def temporaryDemonstrationLogs(message, legend='', sleep_duration=0):
        timestamp = propheticus.shared.Utils.getTimeDifference(Utils.demonstration_start, precision=2)
        Utils.DemonstrationLogs.append([timestamp, sleep_duration, legend, message])
        if sleep_duration > 0:
            time.sleep(sleep_duration)
        # print(legend)

    ''''###############################################################################################'''

    '''
    Messages Related Functions
    '''
    @staticmethod
    def resetLoggedMessages():
        Utils.LoggedMessages = {'error': [], 'status': [], 'warning': []}

    @staticmethod
    def printAcknowledgeMessage(new_line=False):
        Utils.printInputMessage('Acknowledge: (Enter)')
        if new_line is True:
            Utils.printNewLine()

    @staticmethod
    def printErrorMessage(message, acknowledge=True, separator='\n'):
        if isinstance(message, list):
            Utils.LoggedMessages['error'] += message
        else:
            Utils.LoggedMessages['error'].append(message)

        Utils.printBoxedMessage('ERROR', message, separator=separator)
        if acknowledge is True:
            Utils.printAcknowledgeMessage(True)

    @staticmethod
    def printFatalMessage(message):
        Utils.LoggedMessages['error'].append(message)
        Utils.printBoxedMessage('FATAL', message)
        Utils.printAcknowledgeMessage()
        exit('FATAL ERROR OCCURRED')

    @staticmethod
    def printWarningMessage(message, acknowledge=False, separator='\n', break_line=True):
        if isinstance(message, list):
            Utils.LoggedMessages['warning'] += message
        else:
            Utils.LoggedMessages['warning'].append(message)

        Utils.printBoxedMessage('WARNING', message, separator=separator, break_line=break_line)
        if acknowledge is True:
            Utils.printAcknowledgeMessage()

    @staticmethod
    def printNoticeMessage(message):
        Utils.printBoxedMessage('NOTICE', message)

    @staticmethod
    def printBoxedMessage(base, Messages, separator='\n', break_line=True):
        char_split = 140

        if not isinstance(Messages, list):
            Messages = [Messages]

        FinalMessages = []
        for index, message in enumerate(Messages):
            _message = (base + ': ' if index == 0 else f'{index}: ') + message
            if len(_message) < char_split:
                spacing = "".join([' ' * int(((char_split - len(_message)) / 2))])
                _message = (spacing + _message + spacing)[:char_split - 1]

            if break_line is True:
                FinalMessages.append(re.sub("(.{" + str(char_split) + "})", "\\1\n", _message, 0, re.DOTALL))
            else:
                FinalMessages.append(_message)

        spacer = "".join(['!' * char_split])
        print(f'\n{spacer}\n' + f"{separator}".join(FinalMessages) + f'\n{spacer}\n')

    @staticmethod
    def printNoteMessage(message, force=False):
        if Config.log_status or force is True:
            _message = f'NOTE: {message}'
            print(''.join(['*'] * (len(_message) + 1)))
            print(_message)
            print(''.join(['*'] * (len(_message) + 1)) + '\n')

    @staticmethod
    def printStatusMessage(message, inline=False, force=False, label='STATUS'):
        if label is not None:
            label += ': '
        else:
            label = ''

        Utils.LoggedMessages['status'].append(message)
        if Config.log_status or force is True:
            print(f'{label}{message}', end=("" if inline is True else None), flush=True)

    @staticmethod
    def printMessage(message, inline=False, force=False):
        if Config.log_status or force is True:
            print(message, end=("" if inline is True else None), flush=True)

    @staticmethod
    def printNewLine(count=0):
        if Config.log_status is True:
            print('\n' * count)

    @staticmethod
    def printInlineStatusMessage(message):
        if Config.log_status:
            print(message, end="", flush=True)

    @staticmethod
    def printLogMessage(message):
        print('LOG: ' + message)

    @staticmethod
    def printTimeLogMessage(message, start, end=None, precision=0, force=False):
        if Config.log_times or force is True:
            if end is None:
                end = time.time()
            duration = round((end - start), 1)
            if duration > (60 * 3):
                Utils.printLogMessage(message + " took " + str(datetime.timedelta(seconds=duration)))
            else:
                Utils.printLogMessage(message + " took %s seconds" % duration)

    @staticmethod
    def printInputMessage(message):
        if not Utils.isMultiProcess():
            print('INPUT: ' + message)
            input_prepend = " >>  "
            if Config.demonstration_mode is True:
                Utils.printMessage(input_prepend, inline=True)
                action_input, label, Legends, Step = Utils.parseDemoStep()
                if Step is not None:
                    for char in action_input:
                        Utils.printMessage(char, inline=True)
                        time.sleep(0.2)

                    Utils.printNewLine()
                    time.sleep(2)
                    return action_input
                else:
                    return input()
            else:
                return input(input_prepend)
        else:
            print('SILENCED: Threading')

    @staticmethod
    def printConfirmationMessage(message=''):
        Mapping = {'y': 'y', '1': 'y', 'n': 'n', '0': 'n'}
        while True:
            choice = Utils.printInputMessage('Confirmation: ' + message + '\nContinue? y : n\n')
            if choice not in Mapping:
                Utils.printErrorMessage('Invalid selection, please try again')
            else:
                break

        return Mapping[choice]

    @staticmethod
    def consoleClear():
        print('\n' * 2)
        print('&' * 140)
        print('&' * 140)
        print('&' * 140)
        print('\n' * 2)
        # os.system('cls')



    ''''###############################################################################################'''



    '''
    GUI Menus Functions
    '''
    @staticmethod
    def generalMenuData(Context, base_menu_name, MenuData, Configurations, print_menu_configurations=False, show_all_option=True, force_choice=True, menu_key=None):
        if menu_key is None:
            menu_key = base_menu_name.lower().replace(' ', '_')

        def _parseGeneralChoice(choice):
            return Utils.parseChoicesSelection(menu_key, base_menu_name, MenuData, choice, Configurations, force_choice=force_choice)

        return propheticus.shared.Utils.menuData(
            Context=Context,
            menu_name=f'{base_menu_name} Menu',
            MenuData=MenuData,
            callback=_parseGeneralChoice,
            Configurations=(Configurations if print_menu_configurations else None),
            show_all_option=show_all_option,
            force_choice=force_choice
        )

    @staticmethod
    def menuData(Context, menu_name, MenuData, callback, Configurations=None, show_all_option=True, force_choice=True):
        Data = Utils.generateMenuSelectionData(Context, MenuData, callback, show_all_option=show_all_option, force_choice=force_choice)
        def _menuData(choice=0):
            Utils.printMenu(menu_name, Data, Configurations, Context)

        return _menuData

    @staticmethod
    def generateMenuSelectionData(Context, Data, callback, show_all_option=True, force_choice=True):
        list_max_digit = len(str(len(Data)))
        MenuData = {}
        for index, value in enumerate(Data):
            MenuData[str(index + 1).zfill(list_max_digit)] = {'name': value, 'callback': callback}

        if show_all_option is True:
            MenuData[str(len(Data) + 1).zfill(list_max_digit)] = {'name': 'All', 'callback': callback}
            MenuData['_regex'] = {'callback': callback}

        if force_choice is False:
            MenuData['_regex'] = {'callback': callback}

        MenuData['-'] = ''
        MenuData['0'] = {'name': 'Back'}
        MenuData['h'] = {'name': 'Help', 'callback': Context.help()}

        return MenuData

    @staticmethod
    def parseChoicesSelection(property_name, property_display_name, ChoicesData, choice, Configurations, show_selection_message=True, force_choice=True):
        Selections = Utils._parseChoicesSelection(ChoicesData, choice, force_choice)
        if Selections is not False:
            Configurations[property_name] = Selections
            if len(Selections) > 0 and show_selection_message is True:
                Utils.printStatusMessage(property_display_name + ' successfully chosen: ' + ','.join(Configurations[property_name]))

        return False if Selections is False else -1

    @staticmethod
    def _parseChoicesSelection(ChoicesData, choice, force_choice=True):
        Selections = []

        if choice.strip() == str(len(ChoicesData) + 1):
            choice = '1-' + str(len(ChoicesData))

        valid = True
        silence_error = False
        if choice.strip() != '':
            for index in choice.split(','):
                for value2 in index.split('-'):
                    if not value2:
                        if not silence_error:
                            Utils.printErrorMessage('Invalid input format! The valid format must be integers separated by , (individual choices) or - (ranges)')
                            valid = False
                            silence_error = True

                    else:
                        if not value2.isdigit():
                            Utils.printErrorMessage('Invalid choice passed, must be an integer: ' + str(value2))
                            valid = False
                        else:
                            _index = int(value2) - 1
                            if _index < 0 or _index > len(ChoicesData):
                                Utils.printErrorMessage('Invalid choice passed: ' + str(_index))
                                valid = False

            if valid is True:
                for value in choice.split(','):
                    Values = value.split('-')
                    if len(Values) > 1:
                        for value2 in range(int(Values[0]) - 1, int(Values[1])):
                            Selections.append(ChoicesData[value2])
                    else:
                        _value = int(Values[0]) - 1
                        Selections.append(ChoicesData[_value])

        elif force_choice is True:
            Utils.printWarningMessage('No selection made')
            valid = False

        return Selections if valid is True else False

    @staticmethod
    def printBreadCrumb(breadcrumb):
        print(''.join(['-'] * (len(breadcrumb) + 1)))
        print(breadcrumb)
        print(''.join(['-'] * (len(breadcrumb) + 1)) + '\n')

    @staticmethod
    @static_vars(Breadcrumb=['Initialize'])
    def printMenu(label, MenuData, Configurations=None, Context=None, custom_message=None):
        if hasattr(propheticus.Config, 'framework_instance_label'):
            Utils.printMenu.Breadcrumb[0] = Config.framework_instance_label

        while True:
            if Configurations is not None:
                CurrentConfigurations = getattr(Context, Configurations)
                truncate = CurrentConfigurations['config_truncate_configurations'] if 'config_truncate_configurations' in CurrentConfigurations else True
                Utils.printCurrentConfigurations(CurrentConfigurations, hide_empty=True, truncate=truncate)

            if custom_message is not None:
                Utils.printStatusMessage(custom_message, label='CUSTOM')

            breadcrumb_separator = ' >> '
            breadcrumb = breadcrumb_separator + breadcrumb_separator.join(Utils.printMenu.Breadcrumb)
            Utils.printBreadCrumb(breadcrumb)

            print(' ---> ' + label)

            Items = collections.OrderedDict(sorted(MenuData.items())).items() if Config.validate_uuid is True else MenuData.items()
            for key, MenuDetails in Items:
                print('') if MenuDetails == '' or key == '_regex' else print(' -----> ' + str(key) + ' - ' + MenuDetails['name'])
            print('')

            while True:
                choice = Utils.printInputMessage('Select an option:')
                if choice not in MenuData and '_regex' not in MenuData:
                    Utils.printErrorMessage('Invalid selection, please try again', acknowledge=False)
                else:
                    break

            Utils.consoleClear()

            if choice.isdigit() and int(choice) == 0:
                if MenuData['0']['name'] not in ['Back', 'Quit']:
                    Utils.printFatalMessage('FATAL: menu choice 0 must be always assigned to Back')

                if 'callback' in MenuData['0']:
                    MenuData['0']['callback'](choice)
                break
            else:
                if choice in MenuData:
                    Utils.printMenu.Breadcrumb.append(MenuData[choice]['name'])

                returned = MenuData[choice if choice in MenuData else '_regex']['callback'](choice)
                if choice in MenuData:
                    del Utils.printMenu.Breadcrumb[-1]

                if returned == -1:
                    break

    ''''###############################################################################################'''




    '''
    Framework Configurations Related Functions
    '''
    @staticmethod
    def printCurrentConfigurations(Configurations, hide_empty=False, truncate=True):
        configurations = Utils.toStringCurrentConfigurations(Configurations, truncate=truncate, hide_empty=hide_empty)
        Utils.printStatusMessage(configurations)

    @staticmethod
    def toStringCurrentConfigurations(Configurations, truncate=True, hide_empty=False):
        max_chars = 100 if truncate is True else 100000000000000000

        Configs = ['Current configurations:']
        for key, value in Utils.getSafeConfigurationsDict(Configurations, hide_empty).items():
            Configs.append('- [' + key + '] >>> ' + (value[:max_chars] + ', ... ' if len(value) > max_chars else value))

        Configs.append('')

        return '\n'.join(Configs)

    @staticmethod
    def getSafeConfigurations(Configurations):
        SafeConfigurations = [key + '=' + value for key, value in Utils.getSafeConfigurationsDict(Configurations).items()]
        return SafeConfigurations

    @staticmethod
    def getSafeConfigurationsDict(Configurations, hide_empty=False):
        JSONConfigurations = {}
        for key, Values in sorted(Configurations.items()):
            if hide_empty and isinstance(Values, (list, numpy.ndarray, dict, str, type(None), bool)) and not Values:
                continue

            JSONConfigurations[key] = json.dumps(Utils._getSafeConfigurationsDict(Values))

        return collections.OrderedDict(sorted(JSONConfigurations.items()))

    @staticmethod
    def _getSafeConfigurationsDict(ConfigValues):
        _ConfigValues = copy.deepcopy(ConfigValues)
        if isinstance(_ConfigValues, (list, numpy.ndarray, dict, range, tuple)):
            if isinstance(_ConfigValues, dict) and 'label' in _ConfigValues:
                _ConfigValues = _ConfigValues['label']
            else:
                iterator = _ConfigValues.items() if isinstance(_ConfigValues, dict) else enumerate(_ConfigValues)
                ShownSublist = {_key: Utils._getSafeConfigurationsDict(value) for _key, value in iterator}
                _ConfigValues = sorted(ShownSublist.items() if isinstance(_ConfigValues, dict) else ShownSublist.values(), key=lambda x: str(x))
                if isinstance(ConfigValues, dict):
                    _ConfigValues = collections.OrderedDict(_ConfigValues)
        elif not isinstance(_ConfigValues, (str, float, int, bool, type(None))):
            Utils.printWarningMessage('Unexpected data type: ' + str(type(_ConfigValues)))
            _ConfigValues = str(_ConfigValues)

        return _ConfigValues

    @staticmethod
    def parseLogStringConfiguration(experiment_configuration, SelectFields=None):
        Configurations = {}
        for config in experiment_configuration.split('\n'):
            if not config[:3] == '- [':
                continue

            key = config[3:config.index(']')]
            if SelectFields is None or key in SelectFields:
                value = config[config.index('>>>') + 4:]
                Configurations[key] = json.loads(value)

        if SelectFields is not None:
            for field in SelectFields:
                if field not in Configurations:
                    Configurations[field] = None

        return Configurations

    @staticmethod
    def getConfigurationsIdentifier(Configurations):
        return Utils.hash(",".join(sorted(Utils.getSafeConfigurations(Configurations))))

    @staticmethod
    def hash(str_encode):
        return hashlib.md5(str.encode(str_encode)).hexdigest()
    ''''###############################################################################################'''

    '''
    API Dynamic Calls Related Functions
    '''
    @staticmethod
    def dynamicAPICall(CallDetails, OverrideArguments=None, seed=None):
        MethodParameters = CallDetails['parameters']
        method_package = CallDetails['package']
        method_callable = CallDetails['callable']

        Parameters = {parameter: ParameterDetails['default'] for parameter, ParameterDetails in MethodParameters.items() if 'default' in ParameterDetails}
        if seed is not None:
            if 'random_state' in MethodParameters:
                Parameters['random_state'] = seed
            elif 'seed' in MethodParameters:
                Parameters['seed'] = seed

        if OverrideArguments:
            for key, value in OverrideArguments.items():
                Parameters[key] = value

        return Utils.dynamicCall(CallDetails, Parameters)

    @staticmethod
    def dynamicCall(CallDetails, Parameters={}):
        package = CallDetails['package']
        callable_func = CallDetails['callable']
        loaded_module = importlib.import_module(package)
        return getattr(loaded_module, callable_func)(**Parameters)

    ''''###############################################################################################'''


    '''
    Parallelization Required Functions 
    '''
    @staticmethod
    def getBestParallelizationLocation(Configurations, BatchContext=False):
        if Config.thread_level != 'best':
            return Config.thread_level

        if BatchContext:
            configurations_grid_search = None
            for Configuration in Configurations:
                if 'config_grid_search' in Configuration and Configuration['config_grid_search'] is True:
                    configurations_grid_search = True
                    break

            config_load_experiment_models = None
            for Configuration in Configurations:
                if 'config_load_experiment_models' in Configuration and Configuration['config_load_experiment_models'] is not None:
                    config_load_experiment_models = True
                    break

            config_grid_search = configurations_grid_search if configurations_grid_search is not None else Config.InitialConfigurations['config_grid_search']
            if config_grid_search is True:
                return Utils.THREAD_LEVEL_ALGORITHM

            ExecutionCounts = {
                Utils.THREAD_LEVEL_BATCH: len(Configurations),
                Utils.THREAD_LEVEL_RUN: Configurations[0]['config_seed_count'] if 'config_seed_count' in Configurations[0] else Config.InitialConfigurations['config_seed_count'],
            }

            all_algorithms_parallelize = True
            for Configuration in Configurations:
                all_algorithms_parallelize = Utils._validateAlgorithmsParallelization(Configuration)
                if not all_algorithms_parallelize:
                    break
        else:
            config_load_experiment_models = Configurations['config_load_experiment_models']

            if Configurations['config_grid_search'] is True:
                return Utils.THREAD_LEVEL_ALGORITHM

            ExecutionCounts = {
                Utils.THREAD_LEVEL_BATCH: 0,
                Utils.THREAD_LEVEL_RUN: Configurations['config_seed_count'],
            }
            all_algorithms_parallelize = Utils._validateAlgorithmsParallelization(Configurations)

        if all_algorithms_parallelize is True and config_load_experiment_models is None:
            parallelization_level = Utils.THREAD_LEVEL_ALGORITHM
        else:
            parallelization_level =  max(ExecutionCounts.items(), key=operator.itemgetter(1))[0]

        return parallelization_level

    @staticmethod
    def _validateAlgorithmsParallelization(AlgorithmConfig):
        if 'proc_classification' in AlgorithmConfig:
            for algorithm in AlgorithmConfig['proc_classification']:
                if   'n_jobs' not in Config.ClassificationAlgorithmsCallDetails[algorithm]['parameters'] and \
                    ('parallel' not in Config.ClassificationAlgorithmsCallDetails[algorithm] or Config.ClassificationAlgorithmsCallDetails[algorithm]['parallel'] is False):
                    return False

        if 'proc_clustering' in AlgorithmConfig:
            for algorithm in AlgorithmConfig['proc_clustering']:
                if   'n_jobs' not in Config.ClusteringAlgorithmsCallDetails[algorithm]['parameters'] and \
                    ('parallel' not in Config.ClusteringAlgorithmsCallDetails[algorithm] or Config.ClusteringAlgorithmsCallDetails[algorithm]['parallel'] is False):
                    return False

        return True

    @staticmethod
    def istarmap(arguments):
        method = arguments[0]
        args = arguments[1]
        return method(*args)

    @staticmethod
    def pool(pool_count, method, arguments):
        if pool_count < 1:
            propheticus.shared.Utils.printFatalMessage(f'Pool count must be greater than 0. Given: {pool_count}. ' + '\n'.join(map(str, inspect.stack())))

        try:
            ThreadConfigs = multiprocessing.shared_memory.ShareableList(name=Config.thread_config_shared_memory_name)
            ThreadConfigs.shm.close()
            ThreadConfigs.shm.unlink()
        except Exception as e:
            pass

        gc.collect()

        SharedThreadConfigs = multiprocessing.shared_memory.ShareableList([Config.framework_instance, Config.thread_level_ if hasattr(Config, 'thread_level_') else None], name=Config.thread_config_shared_memory_name)

        Pool = multiprocessing.Pool(pool_count, maxtasksperchild=Config.pool_maxtasksperchild)
        PoolData = Pool.starmap(method, arguments) if isinstance(arguments[-1], tuple) else Pool.map(method, arguments)

        Pool.close()
        Pool.join()

        SharedThreadConfigs.shm.close()
        # shared_memory causes issues with spawned processes/threads, try to solve later; this was because un unlink was being done at Config?
        SharedThreadConfigs.shm.unlink()

        try:
            a = 0

        except Exception as e:
            pass

        del SharedThreadConfigs

        gc.collect()

        return PoolData

    ''''###############################################################################################'''

