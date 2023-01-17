"""
Contains the code concerned with comparing the experiments
"""
import matplotlib.pyplot as plt
import scipy.stats
import statsmodels.sandbox.stats.multicomp
import statsmodels.stats.diagnostic
import operator
import copy
import openpyxl

# TODO: change xlrd to openpyxl
import xlrd

import hashlib
import pandas
import itertools
import numpy
import os
from PIL import Image
import subprocess
import pathlib
import shutil
import math
import json
import propheticus
import propheticus.shared

import sys
sys.path.append(propheticus.Config.framework_selected_instance_path)
if os.path.isfile(os.path.join(propheticus.Config.framework_selected_instance_path, 'InstanceConfig.py')):
    from InstanceConfig import InstanceConfig as Config
else:
    import Config as Config

class ExperimentsComparison(object):
    def __init__(self, display_visuals):
        self.display_visuals = display_visuals
        self.display_visuals = False

        self.save_items_path = False

    def compareExperiments(self, CompareExperiments, StatisticalDetails, FocusMetrics=False):
        CompareExperiments = [experiment.strip() for experiment in CompareExperiments]
        if len(CompareExperiments) < 1:
            propheticus.shared.Utils.printErrorMessage('At least 1 experiment must be chosen')
            return

        AvailableExperiments = propheticus.shared.Utils.getAvailableExperiments(skip_config_parse=True)

        ClassificationComparisonLog = [['Focused Metrics: '] + (FocusMetrics if FocusMetrics is not False else ['None'])]
        # ClusteringComparisonLog = []

        # NOTE: add fields here if necessary to inclue in the classification logs details
        DescriptiveFields = [
            'config_binary_classification',
            'config_data_split_parameters',
            'config_seed_count',
            'config_sliding_window',
            'config_undersampling_threshold',
            'custom_hash_parameters',
            'datasets',
            'datasets_classes_remap',
            'pre_target',
            'proc_balance_data',
            'proc_balance_data_parameters',
            'proc_classification',
            'proc_classification_algorithms_parameters',
            'proc_reduce_dimensionality'
        ]
        DescriptiveFields = sorted(DescriptiveFields)

        LogDescriptiveFields = copy.deepcopy(DescriptiveFields)
        LogDescriptiveFields.remove('proc_classification')

        # TODO: we must validate that all the experimetns have the same metrics/order!

        Metrics, ClassificationComparison, RunDetails, RunDetailsByRun, ResultsCV, TimesByExperiment, Results = self.parseResults(CompareExperiments)


        '''
        Validate that all the experiments have the same size and a mininum of 30 runs
        '''
        error_shown = False
        min_runs = 30
        execute_statistical_tests = True
        error = reference_runs_length = False
        for experiment, _Algorithms in ResultsCV.items():
            if len(CompareExperiments) < 2 and len(_Algorithms) < 2:
                propheticus.shared.Utils.printErrorMessage('At least 2 configurations must be present')
                return

            for algorithm, _Metrics in _Algorithms.items():
                algorithm_metrics_lentgh = len(_Metrics[list(_Metrics.keys())[0]])
                if algorithm_metrics_lentgh < min_runs:
                    execute_statistical_tests = False

                if algorithm_metrics_lentgh != reference_runs_length and reference_runs_length is not False and not error_shown:
                    propheticus.shared.Utils.printWarningMessage('Not all the statistical tests have the same number of experiments')
                    execute_statistical_tests = False
                    error_shown = True

                reference_runs_length = algorithm_metrics_lentgh

        if execute_statistical_tests is False:
            propheticus.shared.Utils.printWarningMessage('Not all experiments have the required minimum number of runs. No statistical comparison will be done', acknowledge=False)

        if error is True:
            return

        ExperimentsRanking, Ranking = self.generateExperimentsRanking(Results, FocusMetrics)
        ClassificationComparisonLog += ExperimentsRanking
        ClassificationComparisonLog.append(['Hash', 'Experiment', 'Algorithm'] + LogDescriptiveFields + Metrics)

        '''
        Parse results to the intended file structure
        '''
        ExperimentDetails = {}
        for algorithm, Experiments in ClassificationComparison.items():
            for experiment in Experiments:
                DescriptiveFieldsData = self.getDescriptiveFieldsDataByExperimentDetails(RunDetailsByRun[experiment], LogDescriptiveFields)
                ExperimentDetails[f'{experiment} - {algorithm}'] = [AvailableExperiments[experiment]['filename'], experiment, algorithm] + DescriptiveFieldsData + ClassificationComparison[algorithm][experiment]

        if Ranking is not None:
            RankedExperiments = [Rank[0] for Rank in Ranking]
            RankedFiles = list({Rank[0].split(' - ')[0]: 0 for Rank in Ranking}.keys())
            RankedExperimentDetails = sorted(ExperimentDetails.items(), key=lambda pair: RankedExperiments.index(pair[0]))

            RunDetailsDict = {Details[0]: Details for Details in RunDetails[1:]}
            SortedRunDetails = sorted(RunDetailsDict.items(), key=lambda pair: RankedFiles.index(pair[0]))
            RunDetails = [RunDetails[0]] + [Details for experiment, Details in SortedRunDetails]
        else:
            RankedExperimentDetails = ExperimentDetails.items()

        for experiment, Details in RankedExperimentDetails:
            ClassificationComparisonLog.append(Details)

        # NOTE: How to save with color
        # from openpyxl.styles import PatternFill
        # sheet['A1'].fill = PatternFill(bgColor="FFC7CE", fill_type="solid")

        max_report_records = 20
        if len(CompareExperiments) < max_report_records:
            DimensionalityReductionReport = self.parseDimensionalityReductionLog(ClassificationComparison, RunDetailsByRun, DescriptiveFields)

            if execute_statistical_tests is True:
                StatisticalAnalysisReport = self.performStatisticalAnalysis(StatisticalDetails, ResultsCV)
            else:
                StatisticalAnalysisReport = []
        else:
            propheticus.shared.Utils.printWarningMessage('Too many experiments selected, only ranking will be done')

            StatisticalAnalysisReport = []
            DimensionalityReductionReport = []

        comparisons_hash = hashlib.md5(str.encode(",".join(sorted(CompareExperiments)) + json.dumps(StatisticalDetails) + json.dumps(FocusMetrics))).hexdigest()
        self.save_items_path = os.path.join(Config.framework_instance_generated_comparisons_path, comparisons_hash)
        SheetNames = ['Report', 'Metrics', 'Stats', 'Details']  # 'Dim. Red.',
        propheticus.shared.Utils.saveExcel(
            os.path.join(Config.framework_instance_generated_comparisons_path),
            comparisons_hash + '.Log.xlsx',
            ClassificationComparisonLog,
            [],
            # DimensionalityReductionReport,
            StatisticalAnalysisReport,
            # ClusteringComparisonLog,
            RunDetails,
            SheetNames=SheetNames,
            show_demo=False
        )

        if len(CompareExperiments) < max_report_records:
            self.generateTimeComplexityGraphs(comparisons_hash, TimesByExperiment)
            self.generatePerformanceReportGraphs(comparisons_hash, Results, FocusMetrics)
            self.generatePerformanceGraphs(comparisons_hash, Results, FocusMetrics)

            # TODO: this should be ordered according to ranking

            '''
            Associate relevant experiment images with the comparison file
            '''
            wb_path = os.path.join(Config.framework_instance_generated_comparisons_path, comparisons_hash + '.Log.xlsx')
            Workbook = openpyxl.load_workbook(wb_path)
            Worksheet = Workbook.worksheets[1]
            row = 1

            image_column_increase = 10
            image_row_increase = 25
            column = 1
            Images = {
                'performance_metrics_by_metric': 'Performance Metrics by Metric',
                'performance_report_metrics_by_metric': 'Performance Report Metrics by Metric',
                'performance_metrics_by_experiment': 'Performance Metrics by Experiments',
                'performance_report_metrics_by_experiment': 'Performance Report Metrics by Experiments',
                'time_metrics': 'Time Complexity'
            }

            for image_file_name, image_title in Images.items():
                Worksheet.cell(row=row, column=column).value = image_title
                comparisons_file_path = os.path.join(Config.framework_instance_generated_comparisons_path, comparisons_hash)
                time_metrics_img_path = os.path.join(comparisons_file_path, image_file_name + '.png')
                if os.path.isfile(time_metrics_img_path):
                    img = openpyxl.drawing.image.Image(time_metrics_img_path)
                    Worksheet.add_image(img, openpyxl.utils.cell.get_column_letter(column) + str(row))  # NOTE: (not working though) the reason the image is on the same of the description is to allow filtering in excel
                else:
                    propheticus.shared.Utils.printErrorMessage(image_title + ' - image does not exist for configurations ' + comparisons_hash)

                # column += image_column_increase
                row += image_row_increase

            temp_img_path = None
            Worksheet = Workbook.worksheets[0]
            row = len(ClassificationComparisonLog) + 2
            TempImgs = []

            for _, Details in RankedExperimentDetails:
                experiment = Details[1]
                algorithm = Details[2]
                experiment_file_name = AvailableExperiments[experiment]['filename']
                subfolder = AvailableExperiments[experiment]['subfolder']
                cm_img_filename = os.path.join(experiment_file_name.replace('.Log.xlsx', '') + '.' + algorithm + "_cm.png")
                cm_img_path = os.path.join(Config.framework_instance_generated_classification_path, subfolder, cm_img_filename)
                if os.path.isfile(cm_img_path):
                    pathlib.Path(propheticus.Config.framework_temp_path).mkdir(parents=True, exist_ok=True)
                    temp_img_path = os.path.join(propheticus.Config.framework_temp_path, cm_img_filename)
                    TempImgs.append(temp_img_path)
                    shutil.copyfile(cm_img_path, temp_img_path)

                    img = Image.open(temp_img_path)
                    img_width, img_height = img.size
                    new_height = 480
                    new_width = math.ceil(new_height * img_width / img_height)
                    img = img.resize((new_width, new_height), Image.NEAREST)
                    img.save(temp_img_path)

                    img = openpyxl.drawing.image.Image(temp_img_path)
                    Worksheet.add_image(img, openpyxl.utils.cell.get_column_letter(2) + str(row + 1))
                else:
                    propheticus.shared.Utils.printErrorMessage('Confusion matrix image does not exist for experiment: ' + experiment_file_name.replace('.Log.xlsx', ''), acknowledge=False)

                DescriptiveFieldsData = self.getDescriptiveFieldsDataByExperimentDetails(RunDetailsByRun[experiment], LogDescriptiveFields)
                Details = [experiment_file_name, experiment, algorithm] + DescriptiveFieldsData + ClassificationComparison[algorithm][experiment]
                for index, value in enumerate(Details):
                    Worksheet.cell(row=row, column=(index + 1)).value = value
                row += 26

            comparisons_file_path = os.path.join(Config.framework_instance_generated_comparisons_path, comparisons_hash + '.Log.xlsx')
            Workbook.save(comparisons_file_path)

            for temp_img_path in TempImgs:
                os.remove(temp_img_path)

        if Config.demonstration_mode is True:
            propheticus.shared.Utils.openExternalFileDemo(comparisons_file_path)

        propheticus.shared.Utils.printNewLine()

    def generateExperimentsRanking(self, Results, FocusMetrics):
        RankingLog = [[], ['Ranking According to Scenarios']]

        if FocusMetrics is not False:
            RankingLog.append(['#', 'Experiment'] + FocusMetrics)

            PerformanceByExperimentList = []
            for experiment, Algorithms in Results.items():
                for algorithm, Metrics in Algorithms.items():
                    _Metrics = {key.lower(): value for key, value in Metrics.items()}
                    PerformanceByExperimentList.append([f'{experiment} - {algorithm}'] + numpy.around([_Metrics[i] for i in FocusMetrics], 3).tolist())

            # NOTE: PerformanceByExperimentList is created with the order and metrics defined by the FocusMetrics
            Ranking = sorted(PerformanceByExperimentList, key=operator.itemgetter(*range(1, len(FocusMetrics) + 1)), reverse=True)
            RankingLog += [[index + 1] + Rank for index, Rank in enumerate(Ranking)]
        else:
            RankingLog.append(['No scenario selected'])
            Ranking = None

        return (RankingLog + [[]], Ranking)

    def _openLogs(self, experiment, Details ):
        experiment_file_name = Details['filename']
        subfolder = Details['subfolder']
        wb_path = os.path.join(Config.framework_instance_generated_logs_path, subfolder, experiment_file_name)

        Workbook = openpyxl.load_workbook(wb_path)
        DetailsWorksheet = Workbook.worksheets[0]

        DetailsData = [[DetailsWorksheet.cell(row_index, col_index).value for col_index in range(1, DetailsWorksheet.max_column + 1)] for row_index in range(1, DetailsWorksheet.max_row + 1)]
        ClassificationWorksheet = Workbook.worksheets[1]
        ClassificationData = [[ClassificationWorksheet.cell(row_index, col_index).value for col_index in range(1, ClassificationWorksheet.max_column+1)] for row_index in range(1, ClassificationWorksheet.max_row+1)]

        return experiment, DetailsData, ClassificationData

    def parseResults(self, CompareExperiments):
        RunDetails = [['Experiment', 'Details']]
        RunDetailsByRun = {}
        Metrics = None
        Results = {}
        ResultsCV = {}
        TimesByExperiment = {}
        ClassificationComparison = {}

        AvailableExperiments = propheticus.shared.Utils.getAvailableExperiments(skip_config_parse=True)
        LogsByExp = propheticus.shared.Utils.pool(Config.max_thread_count, self._openLogs, [(experiment, AvailableExperiments[experiment]) for experiment in CompareExperiments])
        propheticus.shared.Utils.printStatusMessage('Parsed all log files')

        for experiment, DetailsData, ClassificationData in LogsByExp:
            Results[experiment] = {}
            ResultsCV[experiment] = {}
            experiment_file_name = AvailableExperiments[experiment]['filename']
            subfolder = AvailableExperiments[experiment]['subfolder']

            RunDetails.append([experiment, DetailsData[0][0]])
            RunDetailsByRun[experiment] = DetailsData[0][0]

            classification_nrows = len(ClassificationData)
            ExpMetrics = [ClassificationData[2][col_index] for col_index in range(1, len(ClassificationData[2]))]

            if Metrics is None:
                Metrics = ExpMetrics

            if ExpMetrics != Metrics:
                propheticus.shared.Utils.printFatalMessage(f'The metrics available in the logs are different across experiments: {Metrics} vs {ExpMetrics} {sorted(set(ExpMetrics).symmetric_difference(set(Metrics)))}')

            for row_index in range(3, len(ClassificationData)):
                cell_value = ClassificationData[row_index][0]
                if cell_value.strip() == '':
                    continue

                algorithm, version = cell_value.strip().split(' - ')
                if version != 'Final':
                    if algorithm not in ResultsCV[experiment]:
                        ResultsCV[experiment][algorithm] = {}

                    for col_index in range(1, len(ClassificationData[row_index])):
                        metric = ClassificationData[2][col_index].lower()
                        value = ClassificationData[row_index][col_index]
                        if metric not in ResultsCV[experiment][algorithm]:
                            ResultsCV[experiment][algorithm][metric] = []
                        ResultsCV[experiment][algorithm][metric].append(value)

                else:
                    if algorithm not in ClassificationComparison:
                        ClassificationComparison[algorithm] = {}

                    if experiment not in ClassificationComparison[algorithm]:
                        ClassificationComparison[algorithm][experiment] = []

                    Results[experiment][algorithm] = {}

                    for col_index in range(1, len(ClassificationData[row_index])):
                        metric = ClassificationData[2][col_index]
                        value = ClassificationData[row_index][col_index]
                        if 'Time' in metric:
                            metric = metric.replace('Time: ', '')
                            if metric not in TimesByExperiment:
                                TimesByExperiment[metric] = {}

                            TimesByExperiment[metric][experiment + '-' + algorithm] = value if value is not None and value.strip() != '' else '-1|-1'
                        Results[experiment][algorithm][metric] = value
                        ClassificationComparison[algorithm][experiment].append(value)

        return Metrics, ClassificationComparison, RunDetails, RunDetailsByRun, ResultsCV, TimesByExperiment, Results


    '''
    Generate Graphs
    '''
    def generateTimeComplexityGraphs(self, comparisons_hash, TimesByExperiment):
        TimeMetrics = [time_metric for time_metric, Experiments in TimesByExperiment.items()]
        TimeExperiments = [experiment for experiment in TimesByExperiment[TimeMetrics[0]].keys()]

        TimeComplexityData = [[float(value.split('|')[0]) for experiment, value in Experiments.items()] for time_metric, Experiments in TimesByExperiment.items()]
        df = pandas.DataFrame(TimeComplexityData, columns=TimeExperiments)
        ax = df.plot.bar(color=["#" + propheticus.shared.Utils.ColourValues[i] for i in range(len(TimeMetrics))])
        ax.set_xticklabels(TimeMetrics)
        ax.legend(loc='lower left', bbox_to_anchor=(0, 0))
        plt.xticks(rotation=40)

        rects = ax.patches
        for rect, label in zip(rects, TimeComplexityData):
            height = rect.get_height()
            ax.text(rect.get_x() + rect.get_width() / 2, height, label, ha='center', va='bottom')

        plt.title('Time Complexity by Experiments')

        propheticus.shared.Utils.saveImage(self.save_items_path, 'time_metrics.png')

        if self.display_visuals is True:
            propheticus.shared.Utils.showImage()

        plt.close()

    def generatePerformanceReportGraphs(self, comparisons_hash, Results, FocusMetrics):
        """
        Generate Report Metrics Comparison Graphs by Experiment

        Parameters
        ----------
        comparisons_hash
        Results
        FocusMetrics

        Returns
        -------

        """
        # Results[experiment][algorithm][ClassificationWorksheet.cell(2, col_index).value] = value
        PerformanceMetrics = []
        MetricsExperiments = [experiment + '-' + algorithm for experiment, Algorithms in Results.items() for algorithm in Algorithms.keys()]
        PerformanceMetricsData = []
        PerformanceMetricsDataByMetric = {}

        # TODO: this must somehow be more generalized; eventually consider different plots, for straigth metrics and report ones?
        IgnoreMetrics = ['']
        for experiment, Algorithms in Results.items():
            for algorithm, Metrics in Algorithms.items():
                PerformanceExperimentMetricsData = []
                for metric, value in Metrics.items():
                    if 'Report' not in metric:
                        continue

                    Performances = value.split('\n')
                    if metric == 'Classification Report':
                        PerfMetrics = Performances[0].split()
                        # Performances = Performances[2:-3]
                        Performances = [performance.split() for performance in Performances if performance.strip() !='' and propheticus.shared.Utils.isInt(performance.split()[0])]
                        for i in range(1, len(PerfMetrics)):
                            for performance in Performances:
                                if FocusMetrics is not False and PerfMetrics[i - 1].lower() not in FocusMetrics:
                                    continue

                                perf_class = performance[0]
                                if perf_class == '0':  # TODO: fix this, currently removing all the control (0) results
                                    continue

                                perf_value = performance[i].strip()
                                perf_value = round(float(perf_value), 2)

                                _metric = PerfMetrics[i - 1] + ' ' + perf_class
                                if _metric not in PerformanceMetrics:
                                    PerformanceMetrics.append(_metric)

                                if _metric not in PerformanceMetricsDataByMetric:
                                    PerformanceMetricsDataByMetric[_metric] = []

                                PerformanceMetricsDataByMetric[_metric].append(perf_value)
                                PerformanceExperimentMetricsData.append(perf_value)
                    else:
                        perf_metric = metric[:metric.index('Report')]
                        if FocusMetrics is not False and perf_metric.lower() not in FocusMetrics:
                            continue

                        for performance in Performances:
                            perf_class = performance.split(':')[0].strip()
                            if perf_class == '0':  # TODO: fix this, currently removing all the control (0) results
                                continue

                            perf_metric += perf_class
                            perf_value = performance.split(':')[1].strip()
                            perf_value = round(float(perf_value), 2)
                            if perf_metric not in PerformanceMetrics:
                                PerformanceMetrics.append(perf_metric)

                            if perf_metric not in PerformanceMetricsDataByMetric:
                                PerformanceMetricsDataByMetric[perf_metric] = []

                            PerformanceMetricsDataByMetric[perf_metric].append(perf_value)
                            PerformanceExperimentMetricsData.append(perf_value)

                #
                #     if isinstance(value, str):
                #         # PerformanceExperimentMetricsData.append(0)
                #         a = 0
                #     else:
                #         PerformanceExperimentMetricsData.append(float(value))
                #         if metric not in PerformanceMetrics:
                #             Perf3ormanceMetrics.append(metric)
                #
                PerformanceMetricsData.append(PerformanceExperimentMetricsData)

        df = pandas.DataFrame(PerformanceMetricsData, columns=PerformanceMetrics)
        ax = df.plot.bar(color=["#" + propheticus.shared.Utils.ColourValues[i] for i in range(len(PerformanceMetrics))])
        plt.xlim(-0.5, len(df) - .5)
        ax.set_xticklabels(MetricsExperiments)
        ax.legend(loc='lower left', bbox_to_anchor=(0, 0))
        plt.xticks(rotation=40)

        if FocusMetrics is not False:
            rects = ax.patches
            for rect, label in zip(rects, PerformanceMetricsData):
                height = rect.get_height()
                ax.text(rect.get_x() + rect.get_width() / 2, height, label, va='bottom')

        plt.title('Performance Report Metrics by Experiments')

        propheticus.shared.Utils.saveImage(self.save_items_path, 'performance_report_metrics_by_experiment.png')

        if self.display_visuals is True:
            propheticus.shared.Utils.showImage()

        plt.close()

        '''
        Experiences Performance by Metrics
        '''
        PerformanceMetricsDataValues = [values for key, values in PerformanceMetricsDataByMetric.items()]
        df = pandas.DataFrame(PerformanceMetricsDataValues, columns=MetricsExperiments)
        ax = df.plot.bar(color=["#" + propheticus.shared.Utils.ColourValues[i] for i in range(len(MetricsExperiments))])
        plt.xlim(-0.5, len(df) - .5)
        ax.set_xticklabels(PerformanceMetrics)
        ax.legend(loc='lower left', bbox_to_anchor=(0, 0))
        plt.xticks(rotation=40)

        if FocusMetrics is not False:
            rects = ax.patches
            for rect, label in zip(rects, PerformanceMetricsDataValues):
                height = rect.get_height()
                ax.text(rect.get_x() + rect.get_width() / 2, height, label, va='bottom')

        plt.title('Performance Report Metrics by Metric')

        propheticus.shared.Utils.saveImage(self.save_items_path, 'performance_report_metrics_by_metric.png')

        if self.display_visuals is True:
            propheticus.shared.Utils.showImage()

        plt.close()

    def generatePerformanceGraphs(self, comparisons_hash, Results, FocusMetrics):
        """
        Generate Metrics Comparison Graphs by Experiment

        Parameters
        ----------
        comparisons_hash
        Results
        FocusMetrics

        Returns
        -------

        """
        # Results[experiment][algorithm][ClassificationWorksheet.cell(2, col_index).value] = value
        PerformanceMetrics = []
        MetricsExperiments = [experiment + '_' + algorithm for experiment, Algorithms in Results.items() for algorithm in Algorithms.keys()]
        PerformanceMetricsData = []
        PerformanceMetricsDataByMetric = {}

        # TODO: this must somehow be more generalized; eventually consider different plots, for straigth metrics and report ones?
        IgnoreMetrics = ['']
        for experiment, Algorithms in Results.items():
            for algorithm, Metrics in Algorithms.items():
                PerformanceExperimentMetricsData = []
                for metric, value in Metrics.items():
                    if FocusMetrics is not False and metric.lower() not in FocusMetrics:
                        continue

                    if value is not None and not isinstance(value, str):
                        value = round(float(value), 2)
                        PerformanceExperimentMetricsData.append(value)
                        if metric not in PerformanceMetrics:
                            PerformanceMetrics.append(metric)

                        if metric not in PerformanceMetricsDataByMetric:
                            PerformanceMetricsDataByMetric[metric] = []

                        PerformanceMetricsDataByMetric[metric].append(value)

                PerformanceMetricsData.append(PerformanceExperimentMetricsData)

        df = pandas.DataFrame(PerformanceMetricsData, columns=PerformanceMetrics)
        ax = df.plot.bar(color=["#" + propheticus.shared.Utils.ColourValues[i] for i in range(len(PerformanceMetrics))])
        ax.set_xticklabels(MetricsExperiments)
        ax.legend(loc='lower left', bbox_to_anchor=(0, 0))
        plt.xlim(-0.5, len(df) - .5)
        plt.xticks(rotation=40)

        if FocusMetrics is not False:
            rects = ax.patches
            for rect, label in zip(rects, PerformanceMetricsData):
                height = rect.get_height()
                ax.text(rect.get_x() + rect.get_width() / 2, height, label, va='bottom')

        plt.title('Performance Metrics by Experiments')

        propheticus.shared.Utils.saveImage(self.save_items_path, 'performance_metrics_by_experiment.png')

        if self.display_visuals is True:
            propheticus.shared.Utils.showImage()

        plt.close()

        '''
        Performance Metrics by Metric
        '''
        PerformanceMetricsDataValues = [values for key, values in PerformanceMetricsDataByMetric.items()]
        df = pandas.DataFrame(PerformanceMetricsDataValues, columns=MetricsExperiments)
        ax = df.plot.bar(color=["#" + propheticus.shared.Utils.ColourValues[i] for i in range(len(MetricsExperiments))])
        ax.set_xticklabels(PerformanceMetrics)
        ax.legend(loc='lower left', bbox_to_anchor=(0, 0))
        plt.xlim(-0.5, len(df) - .5)
        plt.xticks(rotation=40)

        if FocusMetrics is not False:
            rects = ax.patches
            for rect, label in zip(rects, PerformanceMetricsDataValues):
                height = rect.get_height()
                ax.text(rect.get_x() + rect.get_width() / 2, height, label, va='bottom')

        plt.title('Performance Metrics by Metric')

        propheticus.shared.Utils.saveImage(self.save_items_path, 'performance_metrics_by_metric.png')

        if self.display_visuals is True:
            propheticus.shared.Utils.showImage()

        plt.close()


    '''
    Parse Dimensionality Reduction
    '''
    def parseDimensionalityReductionLog(self, ClassificationComparison, RunDetailsByRun, DescriptiveFields):
        AvailableExperiments = propheticus.shared.Utils.getAvailableExperiments(skip_config_parse=True)
        DimensionalityReductionLog = []
        DimensionalityReductions = {}
        for algorithm, Experiments in ClassificationComparison.items():
            for experiment in Experiments:
                experiment_file_name = AvailableExperiments[experiment]['filename']
                subfolder = AvailableExperiments[experiment]['subfolder']
                dimensionality_reduction_path = os.path.join(Config.framework_instance_generated_logs_path, subfolder, experiment_file_name.replace('.Log.', '.dimensionality_reduction.data.'))
                if os.path.isfile(dimensionality_reduction_path) and experiment_file_name not in DimensionalityReductions:
                    # TODO: this code only accounts for 1 dimensionality reduction method; should be expanded to support multiple

                    DRWorkbook = openpyxl.load_workbook(dimensionality_reduction_path)
                    DRSheet = DRWorkbook.worksheets[0]

                    for fr_row in range(1, DRSheet.max_row + 1):
                        if 'Removed Features' in DRSheet.cell(fr_row, 1).value:
                            break

                    DimensionalityReductions[experiment_file_name] = [experiment_file_name, self.getDescriptiveFieldsByExperimentDetails(RunDetailsByRun[experiment], DescriptiveFields), DRSheet.cell(fr_row, 0).value] + DRSheet.row_values(fr_row + 1)

        for Row in list(itertools.zip_longest(*[Data for key, Data in DimensionalityReductions.items()])):
            DimensionalityReductionLog.append(Row)

        return DimensionalityReductionLog

    '''
    propheticus.shared.Utils
    '''
    def getDescriptiveFieldsByExperimentDetails(self, details, DescriptiveFields):
        # TODO: IMPORTANT - REMOVE FOLLOWING HACK, it is just for comparison compatibility
        # return details

        ParsedConfigurations = propheticus.shared.Utils.parseLogStringConfiguration(details, DescriptiveFields)
        return '\n'.join(propheticus.shared.Utils.getSafeConfigurations(ParsedConfigurations))

    def getDescriptiveFieldsDataByExperimentDetails(self, details, DescriptiveFields):
        # TODO: IMPORTANT - REMOVE FOLLOWING HACK, it is just for comparison compatibility
        # return list(range(len(DescriptiveFields)))

        ParsedConfigurations = propheticus.shared.Utils.parseLogStringConfiguration(details, DescriptiveFields)
        return list(propheticus.shared.Utils.getSafeConfigurationsDict(ParsedConfigurations).values())

    '''
    Statistical Analysis
    '''
    def performStatisticalAnalysis(self, StatisticalDetails, ResultsCV):
        metric = StatisticalDetails['metric']

        ComparisonLabels = []
        ComparisonData = []
        for experiment, Algorithms in ResultsCV.items():
            for algorithm, Metrics in Algorithms.items():
                if metric not in Metrics:
                    propheticus.shared.Utils.printFatalMessage('Given metric is not present in the logged metrics: ' + metric + ' . Available metrics are: ' + ', '.join(list(Metrics.keys())))

                ComparisonLabels.append(experiment + ' - ' + algorithm)
                ComparisonData.append(Metrics[metric])

        return self._performStatisticalAnalysis(StatisticalDetails, ComparisonLabels, ComparisonData)

    def _performStatisticalAnalysis(self, StatisticalDetails, ComparisonLabels, ComparisonData):
        confidence_level = StatisticalDetails['confidence_level']
        paired = StatisticalDetails['paired']
        correction_method = StatisticalDetails['correction_method']

        ComparisonData = numpy.around(ComparisonData, 5)

        StatisticalAnalaysisSheet = [['Configurations']] + [['', key, value] for key, value in StatisticalDetails.items()]

        parametric = True

        NormalityTests = {
            'Kol. Smir. Lilliefors': self.testNormalityLilliefors,
            'Shapiro-Wilk': self.testNormalityShapiroWilk
        }

        StatisticalAnalaysisSheet.append([])
        StatisticalAnalaysisSheet.append(['Normality Tests:'])
        for index, Metrics in enumerate(ComparisonData):
            for test, function in NormalityTests.items():
                statistical, pvalue_pt = function(Metrics)
                if pvalue_pt < confidence_level:
                    parametric = False
                    normality_label = 'Not-Normal'
                else:
                    normality_label = 'Normal'

                StatisticalAnalaysisSheet.append(['', ComparisonLabels[index], test, normality_label])

        StatisticalAnalaysisSheet.append([])
        StatisticalAnalaysisSheet.append(['Homogeneity Tests:'])
        statistic_lv, pvalue_lv = self.levene(ComparisonData)
        if pvalue_lv < confidence_level:
            parametric = False
            homogeneity_label = 'Not-Homogeneous'
        else:
            homogeneity_label = 'Homogeneous'

        StatisticalAnalaysisSheet.append(['', 'Levene', homogeneity_label])

        number_categories = len(ComparisonData)
        statistical, pvalue, test = self.statisticalTest(ComparisonData, number_categories, paired, parametric)

        StatisticalAnalaysisSheet.append([])
        StatisticalAnalaysisSheet.append(['Statistical Tests:'])
        StatisticalAnalaysisSheet.append(['', 'Parametric', parametric])
        StatisticalAnalaysisSheet.append(['', 'Test', test])
        StatisticalAnalaysisSheet.append(['', 'H0 Rejected', bool(pvalue < confidence_level)])
        StatisticalAnalaysisSheet.append(['', 'p-value', pvalue])

        if pvalue < confidence_level and number_categories > 2:
            Results, OriginalPValues, ComparedLabels, test = self.multipleComparison(ComparisonData, ComparisonLabels, paired, parametric, confidence_level, correction_method)
            StatisticalAnalaysisSheet.append([])
            StatisticalAnalaysisSheet.append(['Multiple Comparisons:'])
            StatisticalAnalaysisSheet.append(['Comparison', 'H0 Rejected', 'Test', 'Original p-value', 'Corrected p-value'])

            for index, label in enumerate(ComparedLabels):
                StatisticalAnalaysisSheet.append([label, bool(Results[0][index]), test, OriginalPValues[index], Results[1][index]])

        StatisticalAnalaysisSheet.append([])
        StatisticalAnalaysisSheet.append(['Data'])
        StatisticalAnalaysisSheet.append([label for label in ComparisonLabels])
        StatisticalAnalaysisSheet += numpy.transpose(ComparisonData).tolist()

        return StatisticalAnalaysisSheet

    def multipleComparison(self, ComparisonData, ComparisonLabels, paired, parametric, confidence_level, correction_method):
        """
        `correction_method` parameter can be:

        * ``bonferroni`` : one-step correction
        * ``sidak`` : one-step correction
        * ``holm-sidak`` : step down method using Sidak adjustments
        * ``holm`` : step-down method using Bonferroni adjustments
        * ``simes-hochberg`` : step-up method  (independent)
        * ``hommel`` : closed method based on Simes tests (non-negative)
        * ``fdr_bh`` : Benjamini/Hochberg  (non-negative)
        * ``fdr_by`` : Benjamini/Yekutieli (negative)
        * ``fdr_tsbh`` : two stage fdr correction (non-negative)
        * ``fdr_tsbky`` : two stage fdr correction (non-negative)

        Parameters
        ----------
        ComparisonData
        ComparisonLabels
        paired
        parametric
        confidence_level
        correction_method

        Returns
        -------

        """
        # TODO: improve this logic; use https://pypi.org/project/scikit-posthocs/

        '''
        NOTE
        Use Tukey and Dunn where adequate; for dependent ANOVA (repeated measures) use 
        pairwise corrected dependent T-Test
        '''

        ComparisonsLabels = []
        pValues = []
        for i in range(len(ComparisonData)):
            for f in range(i+1, len(ComparisonData)):
                statstisc, pvalue, test = self.statisticalTest([ComparisonData[i], ComparisonData[f]], 2, paired, parametric)
                pValues.append(pvalue)
                ComparisonsLabels.append(ComparisonLabels[i] + ' -> ' + ComparisonLabels[f])

        return statsmodels.sandbox.stats.multicomp.multipletests(pValues, confidence_level, correction_method), pValues, ComparisonsLabels, test

    def statisticalTest(self, Data, number_categories, paired, parametric):
        """
        Based on Ernesto Costa code

        Parameters
        ----------
        Data
        number_categories
        paired
        parametric

        Returns
        -------

        """
        if number_categories > 2:
            # NOTE: # of categories > 2
            if paired is True:
                # NOTE: paired test
                if parametric is True:
                    # NOTE: parametric test
                    statistical_value, pvalue = self.rep_measures_dep_anova(*Data)
                    test = 'rep_measures_dep_anova'
                else:
                    # NOTE: non-parametric test
                    statistical_value, pvalue = scipy.stats.friedmanchisquare(*Data)
                    test = 'friedmanchisquare'
            else:
                # NOTE: unpaired test
                if parametric is True:
                    # NOTE: parametric test
                    statistical_value, pvalue = scipy.stats.f_oneway(*Data)
                    test = 'f_oneway'
                else:
                    # NOTE: non-parametric test
                    statistical_value, pvalue = scipy.stats.kruskal(*Data)
                    test = 'kruskal'
        else:
            # NOTE: # of categories = 2
            if paired is True:
                # NOTE: paired test
                if parametric is True:
                    # NOTE: parametric test
                    statistical_value, pvalue = scipy.stats.ttest_rel(*Data)
                    test = 'ttest_rel'
                else:
                    # NOTE: non-parametric test
                    statistical_value, pvalue = scipy.stats.wilcoxon(*Data)
                    test = 'wilcoxon'
            else:
                # NOTE: unpaired test
                if parametric is True:
                    # NOTE: parametric test
                    eq_pop_var = True
                    # printStatusMessage: validate how to choose the parameter that assumes equal population variance?
                    propheticus.shared.Utils.printWarningMessage('validate how to choose the parameter that assumes equal population variance?')
                    statistical_value, pvalue = scipy.stats.ttest_ind(Data[0], Data[1], equal_var=eq_pop_var)
                    test = 'ttest_ind'
                else:
                    # NOTE: non-parametric test
                    statistical_value, pvalue = scipy.stats.mannwhitneyu(*Data)
                    test = 'mannwhitneyu'

        return statistical_value, pvalue, test

    @staticmethod
    def rep_measures_dep_anova(*args):
        """
        Repeated measures One-Way ANOVA
        From Andy Field, "Discovering Statistics using SPSS (3rd ed.), chapter 13
        March 2016
        Ernesto Costa

        Parameters
        ----------
        args

        Returns
        -------

        """
        df = pandas.DataFrame({index: data for index, data in enumerate(args)})
        grand_mean = df.values.mean()
        # grand_variance = df.values.var(ddof=1)

        row_means = df.mean(axis=1)
        column_means = df.mean(axis=0)

        # n = number of subjects; k = number of conditions/treatments
        n, k = len(df.axes[0]), len(df.axes[1])
        # total number of measurements
        N = df.size  # or n * k

        # degrees of freedom
        df_total = N - 1
        df_between = k - 1
        df_subject = n - 1
        df_within = df_total - df_between
        df_error = df_within - df_subject

        # compute variances
        SS_between = sum(n * [(m - grand_mean) ** 2 for m in column_means])
        SS_within = sum(sum([(df[col] - column_means[i]) ** 2 for i, col in enumerate(df)]))
        SS_subject = sum(k * [(m - grand_mean) ** 2 for m in row_means])
        SS_error = SS_within - SS_subject
        # SS_total = SS_between + SS_within

        # Compute Averages
        MS_between = SS_between / df_between
        MS_error = SS_error / df_error
        MS_subject = SS_subject / df_subject

        # F Statistics
        F = MS_between / MS_error
        # p-value
        p_value = scipy.stats.f.sf(F, df_between, df_error)

        return (F, p_value)

    def testNormalityKolmogorovSmirnov(self, data):
        """
        Ernesto
        Data normalized according to the Central Theorem, where the standard deviation is divided by the square root of the number of samples
        Kolgomorov-Smirnov

        Parameters
        ----------
        data

        Returns
        -------

        """
        norm_data = (data - numpy.mean(data)) / (numpy.std(data) / numpy.sqrt(len(data)))
        return scipy.stats.kstest(norm_data, 'norm')

    def testNormalityLilliefors(self, data):
        """
        Data normalized according to the Central Theorem, where the standard deviation is divided by the square root of the number of samples
        Kolgomorov-Smirnov

        Parameters
        ----------
        data

        Returns
        -------

        """
        norm_data = (data - numpy.mean(data)) / (numpy.std(data) / numpy.sqrt(len(data)))
        return statsmodels.stats.diagnostic.lilliefors(norm_data, 'norm')

    def testNormalityShapiroWilk(self, data):
        """
        Ernesto
        Data normalized according to the Central Theorem, where the standard deviation is divided by the square root of the number of samples
        Shapiro-Wilk

        Parameters
        ----------
        data

        Returns
        -------

        """
        norm_data = (data - numpy.mean(data)) / (numpy.std(data) / numpy.sqrt(len(data)))
        return scipy.stats.shapiro(norm_data)

    def levene(self, data):
        """
        Ernesto
        Test of equal variance.

        Parameters
        ----------
        data

        Returns
        -------

        """
        W, pval = scipy.stats.levene(*data)
        return (W, pval)
